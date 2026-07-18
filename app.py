import streamlit as st
import json
import numpy as np
import colour
import itertools
import pandas as pd
from scipy.optimize import minimize
import base64
import os
import tempfile
from io import BytesIO
from datetime import datetime
# ====== 여기에 아래 두 줄을 추가해 주세요 ======
import openpyxl
from openpyxl.styles import PatternFill
from copy import copy
# ==========================================

# ==========================================
# 0. 설정 및 세션 상태 초기화
# ==========================================
TEMPLATE_FILE = "template.xlsx" 

if "selected_dyes" not in st.session_state:
    st.session_state.selected_dyes = []
if "top_results" not in st.session_state:
    st.session_state.top_results = None
# QTX 파일명과 색상 임시 저장소
if "qtx_filename" not in st.session_state:
    st.session_state.qtx_filename = ""
if "qtx_excel_color" not in st.session_state:
    st.session_state.qtx_excel_color = 14211288 # 기본 회색
# 다운로드용 엑셀 파일 임시 저장소
if "final_excel_bytes" not in st.session_state:
    st.session_state.final_excel_bytes = None
if "final_excel_filename" not in st.session_state:
    st.session_state.final_excel_filename = ""
if "accumulated_data" not in st.session_state:
    st.session_state.accumulated_data = []

def toggle_dye(raw_name):
    if raw_name in st.session_state.selected_dyes:
        st.session_state.selected_dyes.remove(raw_name)
    else:
        st.session_state.selected_dyes.append(raw_name)

def clear_dyes():
    st.session_state.selected_dyes = []
    st.session_state.final_excel_bytes = None # 염료 초기화 시 다운로드 파일도 초기화

# ==========================================
# 1. 엑셀 연동 함수 모음
# ==========================================
def fill_solution_excel(template_file_bytes, all_extracted_data, user_inputs):
    # 기존 코드에 남아있던 pythoncom 관련 내용을 완전히 제거한 순수 openpyxl 버전입니다.
    wb = openpyxl.load_workbook(template_file_bytes)
    ws = wb.worksheets[0]

    # 공통 입력란 채우기 (Ref, From, To 등)
    for i, ans in enumerate(user_inputs): 
        ws.cell(row=3 + i, column=2).value = ans

    # 원본 블록(10~15행)의 병합된 셀 정보 미리 수집
    merges_to_add = []
    for m_range in ws.merged_cells.ranges:
        if 10 <= m_range.min_row <= 15:
            merges_to_add.append((m_range.min_row, m_range.max_row, m_range.min_col, m_range.max_col))

    # [블록 복사 함수] 서식, 높이 등을 그대로 복사
    def copy_block(src_min_row, src_max_row, src_min_col, src_max_col, row_offset):
        for row in range(src_min_row, src_max_row + 1):
            ws.row_dimensions[row + row_offset].height = ws.row_dimensions[row].height
            for col in range(src_min_col, src_max_col + 1):
                src_cell = ws.cell(row=row, column=col)
                dst_cell = ws.cell(row=row + row_offset, column=col)
                
                dst_cell.value = src_cell.value
                if src_cell.has_style:
                    dst_cell.font = copy(src_cell.font)
                    dst_cell.border = copy(src_cell.border)
                    dst_cell.fill = copy(src_cell.fill)
                    dst_cell.number_format = copy(src_cell.number_format)
                    dst_cell.protection = copy(src_cell.protection)
                    dst_cell.alignment = copy(src_cell.alignment)

    # 기준이 될 E11, F11 텍스트
    original_e_val = ws.cell(row=11, column=5).value
    original_f_val = ws.cell(row=11, column=6).value
    existing_e_str = str(original_e_val).strip() if original_e_val is not None else ""
    existing_f_str = str(original_f_val).strip() if original_f_val is not None else ""

    # [데이터 처리]
    for index, data_group in enumerate(all_extracted_data):
        row_offset = index * 6  
        start_row = 10 + row_offset

        # 두 번째 데이터부터는 10~15행을 복사해서 붙여넣기
        if index > 0:
            copy_block(10, 15, 1, 9, row_offset)
            
            # 병합된 셀 구조도 똑같이 복사
            for (m_min_row, m_max_row, m_min_col, m_max_col) in merges_to_add:
                ws.merge_cells(start_row=m_min_row + row_offset, end_row=m_max_row + row_offset, 
                               start_column=m_min_col, end_column=m_max_col)

        # 텍스트 정보 기입
        ws.cell(row=10 + row_offset, column=2).value = data_group['color_name'] # B10
        ws.cell(row=12 + row_offset, column=2).value = data_group['option_letter'] # B12

        # E열, F열 (광원 및 메타머리즘)
        if existing_e_str: 
            ws.cell(row=11 + row_offset, column=5).value = f"{existing_e_str}\n{data_group['light1']}"
        else: 
            ws.cell(row=11 + row_offset, column=5).value = data_group['light1']
        ws.cell(row=12 + row_offset, column=5).value = float(data_group['de_cmc'])
            
        if data_group['light2']:
            if existing_f_str: 
                ws.cell(row=11 + row_offset, column=6).value = f"{existing_f_str}\n{data_group['light2']}"
            else: 
                ws.cell(row=11 + row_offset, column=6).value = data_group['light2']
            ws.cell(row=12 + row_offset, column=6).value = float(data_group['metamerism'])
        else:
            ws.cell(row=11 + row_offset, column=6).value = existing_f_str
            ws.cell(row=12 + row_offset, column=6).value = "-"
        
        # 염료 텍스트 채우기
        num_dyes = len(data_group["dyes"])
        for i in range(num_dyes):
            dye = data_group["dyes"][i]
            ws.cell(row=12 + row_offset + i, column=3).value = dye['dye_name'] # C열
            ws.cell(row=12 + row_offset + i, column=4).value = float(dye['value']) # D열

        # [색상 칠하기 - 도형 대신 A12 셀의 배경색을 채움]
        int_color = data_group["excel_color"]
        # 기존 정수 형태의 RGB를 분해 (r + g*256 + b*65536)
        r = int_color % 256
        g = (int_color // 256) % 256
        b = (int_color // 65536) % 256
        # openpyxl용 Hex 코드로 변환 (ARGB 형식)
        hex_color = f"FF{r:02X}{g:02X}{b:02X}" 
        
        target_cell = ws.cell(row=12 + row_offset, column=1) # A12 기준
        target_cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        target_cell.value = "" 

    # 메모리에 엑셀 파일 저장 후 반환
    final_bytes = BytesIO()
    wb.save(final_bytes)
    final_bytes.seek(0)
    
    return final_bytes


# ==========================================
# 2. 데이터 및 광원 로드
# ==========================================
@st.cache_data
def load_dye_data():
    with open('dye_data.json', 'r') as f:
        raw_data = json.load(f)
        # 💡 수정됨: name 부분에 .strip()을 추가해서 JSON 데이터의 이름표에서도 띄어쓰기를 지웁니다.
        valid_dye_db = {name.strip(): concs for name, concs in raw_data.items() if len(concs) > 0}
        return valid_dye_db

@st.cache_data
def load_dye_mapping(_valid_keys):
    try:
        df = pd.read_excel('dye_list.xlsx', header=None)
        mapping_list = []
        disp_dict = {}
        missing_dyes = [] # 💡 누락된 염료를 추적할 리스트 추가

        for _, row in df.iterrows():
            raw_name = str(row[0]).strip()
            display_name = str(row[1]).strip()
            
            if raw_name in _valid_keys:
                mapping_list.append((raw_name, display_name))
                disp_dict[raw_name] = display_name
            else:
                missing_dyes.append(raw_name) # 💡 유효한 데이터가 없으면 누락 리스트에 추가

        return mapping_list, disp_dict, missing_dyes
    except Exception as e:
        st.error(f"엑셀 파일(dye_list.xlsx)을 읽는 중 오류가 발생했습니다: {e}")
        default_list = [(k, k) for k in sorted(list(_valid_keys))]
        return default_list, {k: k for k in _valid_keys}, []

dye_db = load_dye_data()
# 리턴 받는 값에 missing_dyes 추가
all_dyes_ordered, display_name_dict, missing_dyes = load_dye_mapping(dye_db.keys())

datacolor_tl84_vals = [
    0.91, 0.63, 0.46, 0.37, 1.29, 12.68, 1.59, 1.79, 2.46, 3.38, 
    4.49, 33.94, 12.13, 6.95, 7.19, 7.12, 6.72, 6.13, 5.46, 4.79, 
    5.66, 14.29, 14.96, 8.97, 4.72, 2.33, 1.47, 1.10, 0.89, 0.83, 
    1.18, 4.90, 39.59, 72.84, 32.61, 7.52, 2.83, 1.96, 1.67, 4.43, 
    11.28, 14.76, 12.73, 9.74, 7.33, 9.72, 55.27, 42.58, 13.18, 13.16, 
    12.26, 5.11, 2.07, 2.34, 3.58, 3.01, 2.48, 2.14, 1.54, 1.33, 
    1.46, 1.94, 2.00, 1.20, 1.35, 4.10, 5.58, 2.51, 0.57, 0.27, 
    0.23, 0.21, 0.24, 0.24, 0.20, 0.24, 0.32, 0.26, 0.16, 0.12, 
    0.09
]
wls_tl84 = np.arange(380, 785, 5)
custom_tl84 = colour.SpectralDistribution(dict(zip(wls_tl84, datacolor_tl84_vals)), name='Datacolor_TL84')

LIGHT_MAP = {
    "D65": colour.SDS_ILLUMINANTS.get('D65'),
    "A": colour.SDS_ILLUMINANTS.get('A'),
    "CWF (FL2)": colour.SDS_ILLUMINANTS.get('FL2'),
    "TL84 (FL11)": custom_tl84
}

def get_ks(reflectance):
    return (1 - reflectance)**2 / (2 * reflectance)

def get_ks_normalized(spectrum_map):
    target_wls = np.arange(360, 710, 10)
    sorted_items = sorted(spectrum_map.items(), key=lambda x: int(x[0]))
    existing_wls = np.array([int(k) for k, v in sorted_items])
    existing_vals = np.array([float(v) for k, v in sorted_items])
    normalized_vals = np.interp(target_wls, existing_wls, existing_vals)
    return get_ks(normalized_vals)

blank_r_str = "61.487896,64.536758,67.636276,70.483246,73.516251,75.622711,77.759293,79.583626,80.990044,82.235336,83.458176,84.331772,85.404106,86.164101,86.926323,87.612724,88.086739,88.541801,88.927353,89.348244,89.645943,89.882187,90.113014,90.397278,90.583130,90.746536,90.858932,91.020134,91.199127,91.403587,91.537102,91.670677,91.884819,91.980095,92.083275"
blank_r = np.array([float(x.strip()) / 100.0 for x in blank_r_str.split(',') if x.strip()])
blank_ks = get_ks(blank_r)


# ==========================================
# 3. Streamlit 웹 UI 구성
# ==========================================
st.set_page_config(layout="wide", initial_sidebar_state="expanded", page_title="T/S Colordata", page_icon="logo.png")

st.markdown("""
<style>
    [data-testid="collapsedControl"] { display: none !important; }
    [data-testid="stSidebar"] div.stButton { margin-bottom: -10px; }
</style>
""", unsafe_allow_html=True)


with st.sidebar:
    st.markdown("### 🎨 전체 염료 리스트")
    
    # 💡 데이터가 없어서 누락된 염료가 있다면 사이드바에 안내 메시지 띄우기
    if missing_dyes:
        st.warning(f"⚠️ 데이터 부족으로 제외된 염료 {len(missing_dyes)}개:\n{', '.join(missing_dyes)}")
    
    st.caption("클릭하여 선택 / 해제하세요.")
    
    for raw_name, display_name in all_dyes_ordered:
        btn_type = "primary" if raw_name in st.session_state.selected_dyes else "secondary"
        st.button(
            display_name, 
            key=f"dye_{raw_name}", 
            use_container_width=True, 
            type=btn_type,
            on_click=toggle_dye,
            args=(raw_name,)
        )

col_menu, col_results = st.columns([1.2, 2], gap="large")

with col_menu:
    st.subheader("⚙️ 검색 옵션 설정")
    
    upload_col, color_col = st.columns([2.5, 1])
    with upload_col:
        uploaded_file = st.file_uploader("QTX 파일 업로드", type=['qtx'], label_visibility="collapsed", key="qtx_uploader")
    
    target_r = None
    hex_color = None
    
    if uploaded_file is not None:
        try:
            # 업로드된 파일명 추출 (확장자 제거)
            st.session_state.qtx_filename = os.path.splitext(uploaded_file.name)[0]
            
            content = uploaded_file.getvalue().decode('euc-kr') 
            r_part = content.split("STD_R=")[1].split("\n")[0]
            raw_r_vals = [float(x.strip()) / 100.0 for x in r_part.split(',') if x.strip()]
            
            start_wl = 400 if len(raw_r_vals) == 31 else 360
            for line in content.split('\n'):
                if "STD_REFLLOW=" in line:
                    start_wl = int(line.split("=")[1].replace(',', '').strip())
                    break
            
            current_wls = np.array([start_wl + i * 10 for i in range(len(raw_r_vals))])
            target_wls = np.arange(360, 710, 10)
            target_r = np.interp(target_wls, current_wls, raw_r_vals)
            
            shape_400 = colour.SpectralShape(400, 700, 10)
            wls_400 = np.arange(400, 710, 10)
            viz_wavelengths = dict(zip(wls_400, target_r[4:35])) 
            target_sd_viz = colour.SpectralDistribution(viz_wavelengths).align(shape_400)
            cmfs_viz = colour.MSDS_CMFS['CIE 1964 10 Degree Standard Observer'].copy().align(shape_400)
            ill_viz = LIGHT_MAP["D65"].copy().align(shape_400)
            
            XYZ_viz = colour.sd_to_XYZ(target_sd_viz, cmfs_viz, ill_viz) / 100.0
            white_sd = colour.SpectralDistribution(dict(zip(shape_400.range(), np.ones(len(shape_400))))).align(shape_400)
            wp_D65 = colour.XYZ_to_xy(colour.sd_to_XYZ(white_sd, cmfs_viz, ill_viz) / 100.0)
            RGB_viz = colour.XYZ_to_sRGB(XYZ_viz, illuminant=wp_D65)
            RGB_viz = np.clip(RGB_viz, 0, 1) 
            hex_color = "#{:02x}{:02x}{:02x}".format(int(RGB_viz[0]*255), int(RGB_viz[1]*255), int(RGB_viz[2]*255))
            
            # 엑셀 도형용 RGB 색상값 변환 및 세션 저장
            r_val, g_val, b_val = int(RGB_viz[0]*255), int(RGB_viz[1]*255), int(RGB_viz[2]*255)
            st.session_state.qtx_excel_color = r_val + (g_val * 256) + (b_val * 65536)
                
        except Exception as e:
            st.error(f"QTX 분석 오류: {e}")

    with color_col:
        if hex_color is not None:
            st.markdown(
                f"""
                <div style="display: flex; flex-direction: column; align-items: center; justify-content: center; height: 100%; margin-top: 5px;">
                    <div style="width: 100%; height: 50px; background-color: {hex_color}; border: 1px solid #ccc; border-radius: 8px; box-shadow: 0 1px 3px rgba(0,0,0,0.2);"></div>
                    <div style="font-size: 11px; font-weight: bold; margin-top: 4px; color: #555;">{hex_color.upper()}</div>
                </div>
                """, unsafe_allow_html=True
            )
        else:
            st.markdown(
                f"""
                <div style="display: flex; flex-direction: column; align-items: center; justify-content: center; height: 100%; margin-top: 5px;">
                    <div style="width: 100%; height: 50px; background-color: #f0f2f6; border: 1px dashed #ccc; border-radius: 8px;"></div>
                    <div style="font-size: 11px; margin-top: 4px; color: #999;">미리보기</div>
                </div>
                """, unsafe_allow_html=True
            )

    st.markdown("**💡 광원 우선순위 설정**")
    light_options_all = list(LIGHT_MAP.keys())
    light_options_optional = list(LIGHT_MAP.keys()) + ["없음"]
    none_index = light_options_optional.index("없음")
    
    l_col1, l_col2, l_col3 = st.columns(3)
    light1_name = l_col1.selectbox("1차 광원", light_options_all, index=0)
    light2_name = l_col2.selectbox("2차 광원", light_options_optional, index=none_index) 
    light3_name = l_col3.selectbox("3차 광원", light_options_optional, index=none_index) 

    st.markdown("---")
    st.markdown("**🧪 염료 선택 현황 및 실행**")
    st.markdown(f"선택된 염료: **{len(st.session_state.selected_dyes)}개**")
    
    st.button("🔄 선택 전체 초기화", use_container_width=True, disabled=(len(st.session_state.selected_dyes) == 0), on_click=clear_dyes)

    run_search = False
    if target_r is None:
        st.button("🚀 처방 탐색 시작 (QTX 업로드 필요)", type="primary", use_container_width=True, disabled=True)
    elif len(st.session_state.selected_dyes) < 3:
        st.button("🚀 처방 탐색 시작 (염료 3개 이상 필요)", type="primary", use_container_width=True, disabled=True)
    else:
        run_search = st.button("🚀 처방 탐색 시작", type="primary", use_container_width=True)


with col_results:
    try:
        with open("logo.png", "rb") as image_file:
            logo_base64 = base64.b64encode(image_file.read()).decode()
        st.markdown(
            f"""
            <div style="display: flex; align-items: center; gap: 12px; margin-bottom: 25px;">
                <img src="data:image/png;base64,{logo_base64}" width="45">
                <h2 style="margin: 0; padding: 0;">T/S Colordata</h2>
            </div>
            """, 
            unsafe_allow_html=True
        )
    except:
        st.header("T/S Colordata") # 로고 파일이 없을 때 대비용

    if run_search:
        # 새로운 탐색을 시작하면 이전 다운로드 파일 삭제
        st.session_state.final_excel_bytes = None
        
        selected_pool = st.session_state.selected_dyes
        combos = list(itertools.combinations(selected_pool, 3))
        
        active_lights = [light1_name]
        if light2_name != "없음": active_lights.append(light2_name)
        if light3_name != "없음": active_lights.append(light3_name)
        
        progress_text = "조합을 계산하고 있습니다..."
        my_bar = st.progress(0, text=progress_text)
        results = []
        
        for idx, combo in enumerate(combos):
            combo_display_names = [display_name_dict.get(name, name) for name in combo]
            my_bar.progress((idx + 1) / len(combos), text=f"계산 중: {combo_display_names[0]}, {combo_display_names[1]}, {combo_display_names[2]} ({idx+1}/{len(combos)})")
            
            unit_ks_list = []
            valid_combo = True
            for name in combo:
                available_concs = np.array([float(k) for k in dye_db[name].keys() if float(k) > 0])
                if len(available_concs) == 0:
                    valid_combo = False
                    break
                ref_conc = available_concs[np.argmin(np.abs(available_concs - 1.0))]
                ref_key = [k for k in dye_db[name].keys() if float(k) == ref_conc][0]
                spectrum_map = dye_db[name][ref_key]
                ref_ks_total = get_ks_normalized(spectrum_map)
                ref_ks_net = np.maximum(ref_ks_total - blank_ks, 0)
                unit_ks_list.append(ref_ks_net / ref_conc)
            
            if not valid_combo: continue

            def evaluate_lights_local(conc, return_lab=False):
                total_ks = np.copy(blank_ks)
                for i in range(len(combo)): total_ks += unit_ks_list[i] * conc[i]
                
                est_r = 1 + total_ks - np.sqrt(total_ks**2 + 2 * total_ks)
                shape_5nm = colour.SpectralShape(380, 780, 5)
                start_idx = 4 
                wls_400 = np.arange(400, 710, 10)
                wavelengths = dict(zip(wls_400, est_r[start_idx:35]))
                target_wavelengths = dict(zip(wls_400, target_r[start_idx:35]))
                
                est_sd = colour.SpectralDistribution(wavelengths).align(shape_5nm)
                target_sd = colour.SpectralDistribution(target_wavelengths).align(shape_5nm)
                white_sd = colour.SpectralDistribution(dict(zip(shape_5nm.range(), np.ones(len(shape_5nm.range())))))
                cmfs = colour.MSDS_CMFS['CIE 1964 10 Degree Standard Observer'].copy().align(shape_5nm)
                
                des, labs = [], [] 
                for l_name in active_lights:
                    light_spd = LIGHT_MAP[l_name].copy().align(shape_5nm)
                    wp_XYZ = colour.sd_to_XYZ(white_sd, cmfs, light_spd, method='Integration') / 100.0
                    wp_xy = colour.XYZ_to_xy(wp_XYZ)
                    XYZ_est = colour.sd_to_XYZ(est_sd, cmfs, light_spd, method='Integration') / 100.0
                    XYZ_tgt = colour.sd_to_XYZ(target_sd, cmfs, light_spd, method='Integration') / 100.0
                    lab_est = colour.XYZ_to_Lab(XYZ_est, illuminant=wp_xy)
                    lab_tgt = colour.XYZ_to_Lab(XYZ_tgt, illuminant=wp_xy)
                    de = colour.delta_E(lab_est, lab_tgt, method='CMC', l=2, c=1)
                    des.append(de)
                    labs.append((lab_est, lab_tgt))
                if return_lab: return des, labs
                return des

            def objective_local(conc):
                des, labs = evaluate_lights_local(conc, return_lab=True)
                lab_est, lab_tgt = labs[0] 
                return (lab_est[0] - lab_tgt[0])**2 + (lab_est[1] - lab_tgt[1])**2 + (lab_est[2] - lab_tgt[2])**2

            bnds = [(0.0001, 10), (0.0001, 10), (0.0001, 10)]
            res = minimize(objective_local, x0=[0.1, 0.1, 0.1], bounds=bnds, method='SLSQP', options={'ftol': 1e-6, 'disp': False})
            
            if res.success:
                final_des = evaluate_lights_local(res.x)
                if final_des[0] < 0.5:
                    metamerism_index = sum(final_des[1:]) if len(final_des) > 1 else 0
                    results.append({
                        'combo': combo,
                        'conc': res.x,
                        'des': final_des,
                        'metamerism': metamerism_index,
                        'total_conc': sum(res.x)
                    })

        my_bar.empty() 
        if len(results) > 0:
            results.sort(key=lambda x: x['metamerism'])
            st.session_state.top_results = results[:10]
        else:
            st.session_state.top_results = []
            st.error("⚠️ 유효한 처방을 찾지 못했습니다.")

    # 저장된 결과가 있으면 항상 표시 (새로고침 시에도 유지됨)
    if st.session_state.top_results:
        top_results = st.session_state.top_results
        selected_pool = st.session_state.selected_dyes
        
        row_labels = [f"dE(CMC) {light1_name} (Primary)"]
        if light2_name != "없음": row_labels.append(f"Metamerism {light2_name}")
        if light3_name != "없음": row_labels.append(f"Metamerism {light3_name}")
        row_labels.append("Total concentration [%]")
        row_labels.extend([display_name_dict.get(dye, dye) for dye in selected_pool])
        
        df_dict = {"Property / Dyestuff": row_labels}
        
        for rank, res in enumerate(top_results):
            col_name = f"{rank+1}(3)"
            col_data = []
            
            col_data.append(f"{res['des'][0]:.2f}") 
            
            light_idx = 1
            if light2_name != "없음":
                col_data.append(f"{res['des'][light_idx]:.2f}") 
                light_idx += 1
            if light3_name != "없음":
                col_data.append(f"{res['des'][light_idx]:.2f}") 
                
            col_data.append(f"{res['total_conc']:.4f}") 
            
            for dye in selected_pool:
                if dye in res['combo']:
                    dye_idx = res['combo'].index(dye)
                    col_data.append(f"{res['conc'][dye_idx]:.4f}")
                else:
                    col_data.append("")
            df_dict[col_name] = col_data
        
        df = pd.DataFrame(df_dict)
        df.set_index("Property / Dyestuff", inplace=True)
        
        st.markdown("### 🏆 추천 처방 Top 10 (메타머리즘 최소 순)")
        
        def color_rows(s):
            if s.name.startswith('dE(CMC)'): return ['background-color: #e6f2ff; font-weight: bold'] * len(s)
            elif s.name.startswith('Metamerism'): return ['background-color: #fff9e6; color: #d97706'] * len(s)
            elif s.name == 'Total concentration [%]': return ['background-color: #f3f4f6; font-weight: bold'] * len(s)
            else: return [''] * len(s)
        
        styled_df = df.style.apply(color_rows, axis=1)
        st.dataframe(styled_df, use_container_width=True)
        
        # ==========================================
        # 4. 엑셀 리포트 다운로드 (장바구니 방식 도입)
        # ==========================================
        st.markdown("---")
        st.markdown("### 🛒 엑셀 출력용 데이터 모으기")
        
        available_ranks = list(range(1, len(top_results) + 1))
        selected_rank = st.radio(
            "📌 리스트에 추가할 처방 순위를 선택하세요:", 
            options=available_ranks, 
            horizontal=True
        )
        
        input_color_name = st.text_input("Color Name (색상명):", value=st.session_state.qtx_filename)
        
        # 1. 리스트에 담기 버튼
        if st.button("➕ 현재 처방을 리스트에 추가", use_container_width=True):
            res = top_results[selected_rank - 1]
            dyes_list = []
            for i, dye_raw in enumerate(res['combo']):
                dyes_list.append({
                    "dye_name": display_name_dict.get(dye_raw, dye_raw),
                    "value": round(res['conc'][i], 4)
                })
            
            de_str = f"{res['des'][0]:.2f}"
            meta_str = f"{res['metamerism']:.2f}" if len(res['des']) > 1 else "-"
            light1_short = light1_name.split()[0]
            light2_short = light2_name.split()[0] if light2_name != "없음" else ""
            
            # 장바구니에 넣을 딕셔너리 생성
            new_item = {
                "dyes": dyes_list,
                "de_cmc": de_str,
                "metamerism": meta_str,
                "color_name": input_color_name,
                "option_letter": "R", # 필요시 변경 가능
                "excel_color": st.session_state.qtx_excel_color,
                "light1": light1_short,
                "light2": light2_short
            }
            
            st.session_state.accumulated_data.append(new_item)
            st.success(f"✅ '{input_color_name}' 처방이 리스트에 추가되었습니다! (현재 총 {len(st.session_state.accumulated_data)}개)")

# -----------------------------------------------------------
# 결과창 바깥이나 사이드바, 혹은 하단에 "최종 엑셀 생성 영역" 배치
# (col_results 의 들여쓰기가 끝나는 바깥쪽이나, 적절한 위치에 둡니다)
# -----------------------------------------------------------
st.markdown("---")
st.markdown(f"### 📄 최종 엑셀 리포트 생성 (모인 데이터: {len(st.session_state.accumulated_data)}개)")

if len(st.session_state.accumulated_data) > 0:
    # 담아둔 컬러 이름들 미리보기
    saved_colors = [item['color_name'] for item in st.session_state.accumulated_data]
    st.info(f"**현재 담긴 색상들:** {', '.join(saved_colors)}")
    
    # 엑셀 전체 공통 정보 입력 (한 번만 들어가면 되므로 모아서 빼냅니다)
    with st.expander("📝 리포트 기본 정보 (공통)", expanded=True):
        r_col1, r_col2 = st.columns(2)
        today_date_str = datetime.now().strftime("%d-%b-%y")
        with r_col1:
            input_ref = st.text_input("Ref No. :", value="")
            input_from = st.text_input("From :", value="Ohyoung Inc. /")
        with r_col2:
            input_to = st.text_input("To :", value="Lab Manager")
            input_date = st.text_input("Date :", value=today_date_str)
            input_subject = st.text_input("Subject :", value="Recipe Recommendation")

    col_btn1, col_btn2 = st.columns(2)
    
    with col_btn1:
        if st.button("⚙️ 엑셀 파일 생성하기", type="primary", use_container_width=True):
            if not os.path.exists(TEMPLATE_FILE):
                st.error(f"등록된 템플릿 파일('{TEMPLATE_FILE}')을 찾을 수 없습니다.")
            else:
                with st.spinner("엑셀 파일을 생성하는 중입니다..."):
                    try:
                        with open(TEMPLATE_FILE, "rb") as f:
                            template_bytes = BytesIO(f.read())
                        
                        user_inputs = [input_ref, input_from, input_to, input_date, input_subject]
                        
                        # accumulated_data 전체 리스트를 던져줍니다. (미리 짜두신 반복문이 여기서 빛을 발합니다)
                        final_excel = fill_solution_excel(template_bytes, st.session_state.accumulated_data, user_inputs)
                        
                        st.session_state.final_excel_bytes = final_excel.getvalue()
                        st.session_state.final_excel_filename = f"최종_결과_{datetime.now().strftime('%H%M%S')}.xlsx"
                        st.success("✨ 엑셀 파일이 준비되었습니다!")
                        
                    except Exception as e:
                        st.error(f"엑셀 생성 중 문제가 발생했습니다:\n\n{e}")

    with col_btn2:
        if st.button("🗑️ 리스트 비우기", use_container_width=True):
            st.session_state.accumulated_data = []
            st.session_state.final_excel_bytes = None
            st.rerun()

    # 파일이 생성되었을 때 다운로드 버튼 노출
    if st.session_state.final_excel_bytes is not None:
        st.download_button(
            label="📥 완성된 엑셀 파일 다운로드",
            data=st.session_state.final_excel_bytes,
            file_name=st.session_state.final_excel_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )
else:
    st.write("아직 리스트에 추가된 처방이 없습니다. QTX 파일을 분석하고 처방을 추가해주세요.")